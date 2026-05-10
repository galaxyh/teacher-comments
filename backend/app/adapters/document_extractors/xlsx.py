"""XlsxExtractor — openpyxl with row-cap per sheet.

Per DESIGN-001 §8.2: sheet-per-section, truncate per-sheet at 1000 rows. Output
is GFM pipe tables, one per sheet.
"""

from __future__ import annotations

import asyncio
import io
import logging
from typing import Final

from openpyxl import load_workbook

from app.adapters.document_extractors import ExtractionResult
from app.core.exceptions import DocumentExtractionError, UnsupportedFormatError

logger = logging.getLogger(__name__)

XLSX_MIME: Final[str] = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
ROW_CAP_PER_SHEET: Final[int] = 1000


class XlsxExtractor:
    def supports(self, *, mime_type: str, filename: str) -> bool:
        return mime_type == XLSX_MIME or filename.lower().endswith(".xlsx")

    async def extract(self, *, file_bytes: bytes, filename: str) -> ExtractionResult:
        # `.xls` (legacy OLE) is NOT supported — openpyxl only reads .xlsx (zip-based)
        if file_bytes[:4] == b"\xd0\xcf\x11\xe0":
            raise UnsupportedFormatError(
                "Legacy .xls (OLE) — convert to .xlsx", context={"filename": filename}
            )
        try:
            return await asyncio.to_thread(self._sync_extract, file_bytes, filename)
        except UnsupportedFormatError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise DocumentExtractionError(
                f"Failed to parse .xlsx: {exc.__class__.__name__}: {exc}",
                context={"filename": filename},
            ) from exc

    def _sync_extract(self, file_bytes: bytes, filename: str) -> ExtractionResult:
        # `read_only=True` streams the file (low memory) but blocks formula evaluation,
        # which is fine — we want raw values not computed.
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        warnings: list[str] = []
        sections: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            row_count = 0
            truncated = False
            for raw_row in ws.iter_rows(values_only=True):
                if row_count >= ROW_CAP_PER_SHEET:
                    truncated = True
                    break
                cells = [_cell_to_str(c) for c in raw_row]
                if any(c.strip() for c in cells):  # skip wholly blank rows
                    rows.append(cells)
                    row_count += 1

            if not rows:
                warnings.append(f"sheet '{sheet_name}' was empty")
                continue
            if truncated:
                warnings.append(
                    f"sheet '{sheet_name}' truncated at {ROW_CAP_PER_SHEET} rows"
                )

            sections.append(_render_sheet(sheet_name, rows))

        wb.close()
        text = "\n\n".join(sections).strip()
        return ExtractionResult(
            text=text,
            has_images=False,  # openpyxl images aren't surfaced in our V1 path
            page_count=len(wb.sheetnames),
            warnings=warnings,
        )


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def _render_sheet(name: str, rows: list[list[str]]) -> str:
    """Render a sheet as `## Sheet name` + GFM pipe table.

    First row treated as header even if source had no formatting — Markdown
    pipe tables require a separator line, which has to follow some row.
    """
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = "| " + " | ".join(rows[0]) + " |"
    sep = "| " + " | ".join("---" for _ in range(width)) + " |"
    body = "\n".join("| " + " | ".join(r) + " |" for r in rows[1:])
    return f"## {name}\n{header}\n{sep}\n{body}".strip()
